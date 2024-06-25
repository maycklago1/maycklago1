import customtkinter as Ctk  # Importa a biblioteca customtkinter para criar a interface gráfica
from openpyxl import load_workbook  # Importa a biblioteca openpyxl para manipular arquivos Excel


def buscar_contato(caminho_arquivo='dados.xlsx', nome_planilha='Plan1'):
    # Carrega a planilha Excel do arquivo especificado
    workbook = load_workbook(filename=caminho_arquivo)
    
    # Seleciona a planilha pelo nome fornecido
    sheet = workbook[nome_planilha]
    
    # Inicializa a variável de controle para verificar se o contato foi encontrado
    encontrado = 0
    
    # Converte o nome procurado para maiúsculas
    nome_procurado = f'{nomex[0]}'.upper()
    
    # Itera pelas linhas da planilha, especificamente nas colunas A e B
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
        # Verifica se o valor na coluna A (primeira coluna) corresponde ao nome procurado
        if row[0].value == nome_procurado:
            # Define a variável de controle como 1 (encontrado)
            encontrado = 1
            
            # Obtém o nome e o telefone do contato encontrado
            nome = row[0].value
            telefone = row[1].value
            
            # Atualiza o texto do resultado na interface com o nome e telefone encontrados
            resultado.configure(text=f'{nome} - {telefone}')
            
            # Se o contato foi encontrado, habilita o botão de exclusão
            if encontrado == 1:
                excluir_ctt.configure(state='normal')
    
    # Se o contato não foi encontrado, atualiza o texto do resultado com uma mensagem apropriada
    if encontrado == 0:
        resultado.configure(text='Contato não encontrado na base de dados.')
    
    # Retorna a variável de controle (1 se encontrado, 0 se não)
    return encontrado


def excluir_contato(caminho_arquivo='dados.xlsx', nome_planilha='Plan1'):
    # Carrega a planilha Excel do arquivo especificado
    workbook = load_workbook(filename=caminho_arquivo)

    # Seleciona a planilha pelo nome fornecido
    sheet = workbook[nome_planilha]

    # Converte o nome procurado para maiúsculas
    nome_procurado = f'{nomex[0]}'.upper()
    
    # Itera pelas linhas da planilha, especificamente nas colunas A e B
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
        # Verifica se o valor na coluna A (primeira coluna) corresponde ao nome procurado
        if row[0].value == nome_procurado:
            # Se o nome for encontrado, apaga a linha correspondente
            sheet.delete_rows(row[0].row, 1)
            
            # Atualiza o texto do resultado na interface com uma mensagem de sucesso
            resultado.configure(text='Contato excluído com sucesso!')
            break
    
    # Salva as mudanças feitas na planilha
    workbook.save(filename=caminho_arquivo)


def adicionar_contato(caminho_arquivo='dados.xlsx', nome_planilha='Plan1'):
    # Carrega a planilha Excel do arquivo especificado
    workbook = load_workbook(filename=caminho_arquivo)

    # Seleciona a planilha pelo nome fornecido
    sheet = workbook[nome_planilha]

    # Obtém o nome e o telefone do novo contato a partir dos campos de entrada
    adicionar_nome = add_nome.get().upper()
    adicionar_telefone = add_telefone.get().upper()

    # Adiciona o novo contato como uma nova linha na planilha
    sheet.append([adicionar_nome, adicionar_telefone])

    # Salva as mudanças feitas na planilha
    workbook.save(filename=caminho_arquivo)
    
    # Atualiza o texto do resultado na interface com uma mensagem de sucesso
    resultado_adicionar.configure(text='Contato adicionado com sucesso!')


def button_click_event():
    # Limpa a lista de pesquisa e adiciona o novo nome pesquisado
    nomex.clear()
    nomex.append(campo_pesquisa.get())
    
    # Habilita o botão de limpar pesquisa
    limpar_pesquisa.configure(state='normal')
    
    # Chama a função de busca de contato
    buscar_contato()
    

def limpar_pesq():
    # Limpa o campo de entrada de pesquisa
    campo_pesquisa.delete(0, Ctk.END)
    
    # Redefine o texto do resultado da pesquisa
    resultado.configure(text='Aguardando pesquisa...')
    
    # Limpa a lista de pesquisa
    nomex.clear()
    
    # Desabilita o botão de excluir contato
    excluir_ctt.configure(state='disabled')
        
def limpar_adicionar():
    # Limpa os campos de entrada de adicionar contato
    add_nome.delete(0, Ctk.END)
    add_telefone.delete(0, Ctk.END)
    
    # Redefine o texto do resultado de adicionar contato
    resultado_adicionar.configure(text='Insira os dados do contato para adicionar.')
    

# Variáveis globais para armazenar os dados de contatos e pesquisa
nome = ''
telefone = ''
nomex = []
nome_z = []

# Criação da janela principal da aplicação
janela = Ctk.CTk()

# Configurações da janela principal
janela.title("AGENDA")  # Define o título da janela
janela.geometry("560x720+650+170")  # Define o tamanho e a posição inicial da janela
janela.maxsize(560,720)  # Define o tamanho máximo da janela
janela.minsize(560,720)  # Define o tamanho mínimo da janela
janela.config(bg='#f8f69f')  # Define a cor de fundo da janela
janela.grid_columnconfigure((0, 1), weight=1)  # Configura as colunas para expandirem igualmente


# Cria um rótulo de título na janela principal
titulo = Ctk.CTkLabel(janela, text='AGENDA TELEFÔNICA', font=(
    'Arial Black', 36), text_color=('white'), fg_color=('#000039'))
titulo.grid(row=0, column=0, padx=20, pady=20, sticky="ew", columnspan=2)


# Cria um rótulo para a seção de pesquisa
titulo_pesquisar = Ctk.CTkLabel(janela, text='PESQUISAR', font=(
    'Arial Black', 18), fg_color=('#3e3e53'))
titulo_pesquisar.grid(row=1, column=0, padx=20, pady=20,
                      sticky="ew", columnspan=2)


# Cria um campo de entrada de texto para a pesquisa de contatos
campo_pesquisa = Ctk.CTkEntry(janela,font= ('Lucida Sans Typewriter Negrito Inclinado', 12), text_color='black' , placeholder_text='nome',fg_color='#7c7b6c', placeholder_text_color='black', corner_radius=-10)
campo_pesquisa.grid(row=2, column=0, padx=20, pady=20)

# Cria um botão para iniciar a pesquisa de contatos
pesquisa = Ctk.CTkButton(janela, font=('Copperplate Gothic Bold', 12), text='PESQUISAR',
                         command=button_click_event, corner_radius=-10, fg_color='#000039', hover_color='#042a5f')
pesquisa.grid(row=2, column=1, padx=20, pady=20)

# Cria um rótulo para exibir o resultado da pesquisa
resultado = Ctk.CTkLabel(janela, font=('Lucida Console', 14), text_color='#000039',
                         text='RESULTADO DA PESQUISA', width=150, fg_color='#bab986')
resultado.grid(row=3, column=0, padx=20, pady=20, sticky="ew", columnspan=2)

# Cria um botão para limpar a pesquisa atual
limpar_pesquisa = Ctk.CTkButton(janela, font=('Copperplate Gothic Bold', 12), text='LIMPAR PESQUISA', fg_color='#2aa17b', hover_color='#57c69e', corner_radius=-10, command=limpar_pesq, state='disabled')
limpar_pesquisa.grid(row=4, column=0, padx=20, pady=20)

# Cria um botão para excluir o contato encontrado na pesquisa
excluir_ctt = Ctk.CTkButton(
    janela, font=('Copperplate Gothic Bold', 12), text='EXCLUIR CONTATO!', fg_color='#8c0908', text_color='white', hover_color='#b61c17', corner_radius=-10, state='disabled', command=excluir_contato)
excluir_ctt.grid(row=4, column=1, padx=20, pady=20)


# SEÇÃO DE ADIÇÃO DE CONTATO

# Cria um rótulo para a seção de adicionar um novo contato
titulo_adicionar = Ctk.CTkLabel(janela, text='ADICIONAR NOVO CONTATO', font=(
    'Arial Black', 18), fg_color=('#3e3e53'))
titulo_adicionar.grid(row=5, column=0, padx=20, pady=20,
                      sticky="ew", columnspan=2)

# Cria um campo de entrada de texto para o nome do novo contato
add_nome = Ctk.CTkEntry(janela, font= ('Lucida Sans Typewriter Negrito Inclinado', 12), text_color='black' ,placeholder_text='Nome do contato',
                        fg_color='#7c7b6c', placeholder_text_color='black', corner_radius=-10)
add_nome.grid(row=6, column=0, padx=20, pady=20)

# Cria um campo de entrada de texto para o telefone do novo contato
add_telefone = Ctk.CTkEntry(janela, font= ('Lucida Sans Typewriter Negrito Inclinado', 12), text_color='black' ,placeholder_text='Numero do telefone',
                            fg_color='#7c7b6c', placeholder_text_color='black', corner_radius=-10)
add_telefone.grid(row=7, column=0, padx=20, pady=20)

# Cria um botão para adicionar o novo contato
botao_adicionar = Ctk.CTkButton(janela, font=('Copperplate Gothic Bold', 12),text='ADICIONAR CONTATO', corner_radius=-10, fg_color='#000039', hover_color='#042a5f', command=adicionar_contato)
botao_adicionar.grid(row=6, column=1, padx=20, pady=20)

# Cria um botão para limpar os campos de entrada de adicionar contato
limpar_add = Ctk.CTkButton(janela, font=('Copperplate Gothic Bold', 12), text='LIMPAR CAMPOS', fg_color='#2aa17b', hover_color='#57c69e', corner_radius=-10, command=limpar_adicionar)
limpar_add.grid(row=7, column=1, padx=20, pady=20)

# Cria um rótulo para exibir o status da adição de um novo contato
resultado_adicionar = Ctk.CTkLabel(janela,  font=('Lucida Console', 14), text_color='#000039',
                                   text='Insira os dados do contato.', width=200,  fg_color='#bab986')
resultado_adicionar.grid(row=8, column=0, padx=20,
                         pady=20, sticky="ew", columnspan=2)


# Inicia o loop principal da interface gráfica
janela.mainloop()
